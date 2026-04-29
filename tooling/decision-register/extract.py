#!/usr/bin/env python3
"""Extract CCB decisions from this repository's RFC PRs into the
CCB Decision Register (xlsx) per WI §7.7 / Appendix E.

Reads PR state and labels via the GitHub API. One row per merged or closed
RFC PR. The xlsx output is the artifact that gets uploaded to and signed in
Arena/Windchill (per SOP-00028) — this script is the validated extract.

Usage:
    GITHUB_TOKEN=... python extract.py \\
        --repo pl-buiquang/test-rfcs \\
        --output ccb_decision_register.xlsx \\
        [--since 2026-01-01]
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any, Iterator

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

API = "https://api.github.com"
FRONTMATTER_RE = re.compile(r"\A---\n(.*?)\n---\n", re.DOTALL)

# Map GitHub team slug -> CCB discipline. Replace these placeholders before
# adopting in production. The pipeline is deterministic and these mappings
# are versioned with the repo.
TEAM_TO_DISCIPLINE: dict[str, str] = {
    "qa": "Quality Assurance",
    "regulatory": "Regulatory Affairs",
    "engineering": "Engineering",
    "product": "Product Management",
    "clinical": "Clinical/Medical",
    "architecture-review-board": "Engineering",
}


@dataclass
class DecisionRow:
    rfc_id: str = ""
    title: str = ""
    product: str = ""
    target_version: str = ""
    tier: str = "standard"
    decision: str = ""
    decision_date: str = ""
    approvers: str = ""
    disciplines_covered: str = ""
    impact_summary: str = ""
    rationale: str = ""
    defer_re_eval_conditions: str = ""
    change_control_ref: str = ""
    pr_url: str = ""

    @classmethod
    def columns(cls) -> list[str]:
        return [f.name for f in cls.__dataclass_fields__.values()]


class GitHub:
    def __init__(self, token: str, repo: str):
        self.repo = repo
        self.s = requests.Session()
        self.s.headers.update(
            {
                "Authorization": f"Bearer {token}",
                "Accept": "application/vnd.github+json",
                "X-GitHub-Api-Version": "2022-11-28",
            }
        )

    def _paged(self, url: str, params: dict[str, Any] | None = None) -> Iterator[dict]:
        params = dict(params or {})
        params.setdefault("per_page", 100)
        while url:
            r = self.s.get(url, params=params)
            r.raise_for_status()
            yield from r.json()
            url = r.links.get("next", {}).get("url", "")
            params = None  # next URL already encodes them

    def closed_prs(self) -> Iterator[dict]:
        yield from self._paged(
            f"{API}/repos/{self.repo}/pulls",
            {"state": "closed", "sort": "updated", "direction": "desc"},
        )

    def open_prs(self) -> Iterator[dict]:
        yield from self._paged(f"{API}/repos/{self.repo}/pulls", {"state": "open"})

    def reviews(self, pr_number: int) -> list[dict]:
        return list(self._paged(f"{API}/repos/{self.repo}/pulls/{pr_number}/reviews"))

    def comments(self, pr_number: int) -> list[dict]:
        return list(
            self._paged(f"{API}/repos/{self.repo}/issues/{pr_number}/comments")
        )

    def file(self, ref: str, path: str) -> str | None:
        r = self.s.get(
            f"{API}/repos/{self.repo}/contents/{path}",
            params={"ref": ref},
            headers={"Accept": "application/vnd.github.raw"},
        )
        if r.status_code == 404:
            return None
        r.raise_for_status()
        return r.text

    def user_teams(self, login: str) -> list[str]:
        # Best-effort discipline lookup. Requires read:org scope and the user
        # being an org member. On personal repos this returns []; rationale
        # is the manual fallback.
        try:
            org = self.repo.split("/")[0]
            teams = list(self._paged(f"{API}/orgs/{org}/teams"))
        except requests.HTTPError:
            return []
        slugs: list[str] = []
        for team in teams:
            r = self.s.get(
                f"{API}/orgs/{org}/teams/{team['slug']}/memberships/{login}"
            )
            if r.status_code == 200:
                slugs.append(team["slug"])
        return slugs


def parse_frontmatter(text: str) -> dict[str, Any]:
    m = FRONTMATTER_RE.match(text)
    if not m:
        return {}
    try:
        return yaml.safe_load(m.group(1)) or {}
    except yaml.YAMLError:
        return {}


def find_rfc_file_path(gh: GitHub, pr: dict) -> str | None:
    files = list(
        gh._paged(f"{API}/repos/{gh.repo}/pulls/{pr['number']}/files")
    )
    for f in files:
        if f["filename"].startswith("rfcs/") and f["filename"].endswith(".md"):
            return f["filename"]
    return None


def decision_for(pr: dict) -> str | None:
    labels = {l["name"] for l in pr.get("labels", [])}
    if pr["merged_at"]:
        return "accept"
    if "status:rejected" in labels:
        return "reject"
    if "status:deferred" in labels:
        return "defer"
    if pr["state"] == "closed":
        # Closed without explicit label — caller can decide whether to skip.
        return None
    return None


def discipline_of(gh: GitHub, login: str) -> str:
    for slug in gh.user_teams(login):
        if slug in TEAM_TO_DISCIPLINE:
            return TEAM_TO_DISCIPLINE[slug]
    return "Unknown"


def extract_rationale(comments: list[dict]) -> str:
    # Heuristic: last comment whose body starts with "Decision:" or contains
    # "rationale" wins. Production version should look at PR reviews + the
    # final comment authored by the CCB Chair (tracked via a label).
    for c in reversed(comments):
        body = (c.get("body") or "").strip()
        if body.lower().startswith("decision:") or "rationale" in body.lower():
            return body
    return ""


def build_row(gh: GitHub, pr: dict) -> DecisionRow | None:
    decision = decision_for(pr)
    if decision is None:
        return None

    rfc_path = find_rfc_file_path(gh, pr)
    fm: dict[str, Any] = {}
    if rfc_path:
        ref = pr["merge_commit_sha"] or pr["head"]["sha"]
        text = gh.file(ref, rfc_path) or ""
        fm = parse_frontmatter(text)

    labels = {l["name"] for l in pr.get("labels", [])}
    tier = "lightweight" if "rfc:lightweight" in labels else "standard"
    decision_dt = pr["merged_at"] or pr["closed_at"] or ""

    reviews = gh.reviews(pr["number"])
    approvers: list[str] = []
    seen: set[str] = set()
    disciplines: set[str] = set()
    for rv in reviews:
        if rv.get("state") != "APPROVED":
            continue
        login = rv["user"]["login"]
        if login in seen:
            continue
        seen.add(login)
        d = discipline_of(gh, login)
        disciplines.add(d)
        approvers.append(f"{login} [{d}] @ {rv['submitted_at']}")

    rationale = extract_rationale(gh.comments(pr["number"]))
    defer_conditions = rationale if decision == "defer" else ""

    return DecisionRow(
        rfc_id=str(fm.get("rfc_id", "")).zfill(4) if fm.get("rfc_id") else "",
        title=fm.get("title", pr["title"]),
        product=fm.get("product", ""),
        target_version=fm.get("version", ""),
        tier=tier,
        decision=decision,
        decision_date=decision_dt,
        approvers="; ".join(approvers),
        disciplines_covered="; ".join(sorted(disciplines)),
        impact_summary="(see PR for full Impact Triage)",
        rationale=rationale if decision != "defer" else "",
        defer_re_eval_conditions=defer_conditions,
        change_control_ref=fm.get("change_control", ""),
        pr_url=pr["html_url"],
    )


def write_xlsx(rows: list[DecisionRow], output: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "CCB Decision Register"

    columns = DecisionRow.columns()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4F7F")
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=getattr(row, name))

    for col_idx, name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            14, min(60, len(name) + 4)
        )

    ws.freeze_panes = "A2"
    wb.save(output)


def parse_since(s: str | None) -> datetime | None:
    if not s:
        return None
    return datetime.fromisoformat(s).replace(tzinfo=timezone.utc)


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--repo", required=True, help="owner/repo")
    p.add_argument("--output", default="ccb_decision_register.xlsx")
    p.add_argument(
        "--since",
        help="ISO date (YYYY-MM-DD) — only PRs updated since this date",
    )
    p.add_argument(
        "--include-deferred-open",
        action="store_true",
        help="Also include open PRs with status:deferred (they are not yet closed)",
    )
    args = p.parse_args(argv)

    token = os.environ.get("GITHUB_TOKEN")
    if not token:
        print("GITHUB_TOKEN env var is required.", file=sys.stderr)
        return 2

    since = parse_since(args.since)
    gh = GitHub(token, args.repo)

    rows: list[DecisionRow] = []
    for pr in gh.closed_prs():
        if since:
            updated = datetime.fromisoformat(pr["updated_at"].replace("Z", "+00:00"))
            if updated < since:
                break
        labels = {l["name"] for l in pr.get("labels", [])}
        if "rfc" not in labels:
            continue
        row = build_row(gh, pr)
        if row:
            rows.append(row)

    if args.include_deferred_open:
        for pr in gh.open_prs():
            labels = {l["name"] for l in pr.get("labels", [])}
            if "status:deferred" in labels and "rfc" in labels:
                row = build_row(gh, pr)
                if row:
                    rows.append(row)

    rows.sort(key=lambda r: r.rfc_id)
    write_xlsx(rows, args.output)
    print(f"Wrote {len(rows)} decision row(s) to {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
